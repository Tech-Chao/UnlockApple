#coding=utf-8
import json
import openpyxl
import sys
import os

base_path =  os.getcwd()
excel_path = os.path.join(base_path, "asset", "AppleID.xlsx")

class HandExcel:
    def load_excel(self):
        open_excel = openpyxl.load_workbook(excel_path)
        return open_excel
    def get_sheet_data(self,index=None):
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
            data = self.load_excel()[sheet_name[index]]
            return data
    def get_cell_value(self,row,cols):
        "获取某个单元格的内容"
        value = self.get_sheet_data().cell(row=row,column=cols).value
        return value
    def get_row(self):
        '''获取行数'''
        row = self.get_sheet_data().max_row
        return row
    def get_row_value(self,row):
        '''获取一整行数据'''
        row_list =[]
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list
    def excel_write_data(self,row,cols,value):
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(excel_path)

    def get_columns_value(self,key=None):
        '''获取某一列的数据'''
        column_list = []
        if key ==None:
            key ='A'
        column_list_data = self.get_sheet_data()[key]
        for i in column_list_data:
            column_list.append(i.value)
        return column_list

    def get_rows_number(self,case_id):
        '''获取行号'''
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num +1
        return num

    def get_excel_data(self):
        '''获取excel里面所有的数据'''
        data_list = []
        for i in range(self.get_row()):
            data_list.append(self.get_row_value(i+2))
        return data_list


if __name__ =="__main__":
    handexcel = HandExcel()
    print(handexcel.get_excel_data())